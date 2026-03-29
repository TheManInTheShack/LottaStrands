import openpyxl as xl
import pandas as pd
import os
from time import sleep


# ------------------------------------------------------------------------------
# check if a file is writeable currently
# ------------------------------------------------------------------------------
def can_write_to_excel(book, tell=False):
    if os.path.isfile(book):
        if tell:
            print(f"...checking write permissions for file {book}...")
        try:
            temp = book + "xyz"
            x = os.rename(book, temp)
            x = os.rename(temp, book)
            return True
        except:
            return False
    else:
        return True

# ------------------------------------------------------------------------------
# Write a plain list-of-lists to excel
# ------------------------------------------------------------------------------
def write_matrix_to_excel_sheet(ws, matrix, row_offset=0, col_offset=0):
    # --------------------------------------------------------------------------
    # There has to be one to start with
    # --------------------------------------------------------------------------
    if not matrix:
        return False

    # --------------------------------------------------------------------------
    # If the thing is not a list
    # --------------------------------------------------------------------------
    if not isinstance(matrix, list):
        print("Hey, you're trying to pass something that isn't a matrix (main structure not a list)")
        return False

    # --------------------------------------------------------------------------
    # Write the matrix
    # --------------------------------------------------------------------------
    for i, rec in enumerate(matrix):
        # ------------------------------------------------------------------
        # Maybe it's not a line
        # ------------------------------------------------------------------
        if not (isinstance(rec, list) or isinstance(rec, tuple)):
            print(rec)
            print("Hey, you're trying to pass something that isn't a matrix (a list, but not all of the items in the list are lists)")
            sys.exit()
            return False

        # ------------------------------------------------------------------
        # Maybe there's nothing on the line
        # ------------------------------------------------------------------
        if len(rec) == 0:
            continue

        # ----------------------------------------------------------------------
        # Each cell
        # ----------------------------------------------------------------------
        for j, col in enumerate(rec):
            ws.cell(row=i+1+row_offset, column=j+1+col_offset).value = col

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Make and Write to an excel sheet using a grid of data and a dict of formatting
# ------------------------------------------------------------------------------
def write_excel_sheet(wb, wsname, data, formatting, row_offset=0, col_offset=0, replace=True, cond=True, display=True):
    # --------------------------------------------------------------------------
    # We only do this if the condition is true
    # --------------------------------------------------------------------------
    if not (cond):
        return False

    # --------------------------------------------------------------------------
    # Remove the existing version of the sheet
    # --------------------------------------------------------------------------
    if replace and wsname in wb:
        ws = wb[wsname]
        last_row = len(list(ws.rows))
        ws.delete_rows(1,last_row+1)
    else:
        ws = wb.create_sheet(wsname)

    # --------------------------------------------------------------------------
    # Make the sheet
    # --------------------------------------------------------------------------
    if display:
        print(f"...writing sheet: {wsname}...")

    # --------------------------------------------------------------------------
    # Write in the data
    # --------------------------------------------------------------------------
    w = write_matrix_to_excel_sheet(ws, data, row_offset, col_offset)

    # --------------------------------------------------------------------------
    # Formatting - column widths and row height
    # --------------------------------------------------------------------------
    if 'column_widths' in formatting:
        for i, column_width in enumerate(formatting['column_widths']):
            ws.column_dimensions[xl.utils.get_column_letter(i+1)].width = column_width

    if 'fix_row_height_after' in formatting:
        for i, row in enumerate(data):
            if i < formatting['fix_row_height_after']+1:
                continue
            ws.row_dimensions[i].height = 15

    if 'short_rows' in formatting:
        for i, row in enumerate(formatting['short_rows']):
            ws.row_dimensions[row].height = 5


    # --------------------------------------------------------------------------
    # Formatting - tint block of cells (this happens before the black/white
    # because that one may overwrite this one)
    # --------------------------------------------------------------------------
    if 'user_cols' in formatting:
        for col in formatting['user_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "tint_gray")

    if 'highlight_cols_yellow' in formatting:
        for col in formatting['user_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "tint_yellow")

    if 'highlight_cols_green' in formatting:
        for col in formatting['user_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "tint_green")

    if 'conditional_user_cols' in formatting:
        for cond_set in formatting['conditional_user_cols']:
            for col in cond_set['cols']:
                for i, row in enumerate(data):
                    cond_cell = cond_set['cond'] + str(i+1)

                    if valid_value(ws[cond_cell].value, allow_zero=True) and 'value' in cond_set:
                        if ws[cond_cell].value == cond_set['value']:
                            cell = ws[col + str(i+1)]
                            apply_formatting_to_cell(cell, "tint_gray")

                    elif valid_value(ws[cond_cell].value, allow_zero=True) and 'is_number' in cond_set:
                        if cond_set['is_number']:
                            if is_number(ws[cond_cell].value):
                                cell = ws[col + str(i+1)]
                                apply_formatting_to_cell(cell, "tint_gray")
                        else:
                            if not is_number(ws[cond_cell].value):
                                cell = ws[col + str(i+1)]
                                apply_formatting_to_cell(cell, "tint_gray")

                    else:
                        if ws[cond_cell].value:
                            cell = ws[col + str(i+1)]
                            apply_formatting_to_cell(cell, "tint_gray")

    if 'highlight_rows_yellow' in formatting:
        for rownum in formatting['highlight_rows_yellow']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "tint_yellow")

    if 'highlight_rows_green' in formatting:
        for rownum in formatting['highlight_rows_green']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "tint_green")

    if 'highlight_rows_gray' in formatting:
        for rownum in formatting['highlight_rows_gray']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "tint_gray")

    # --------------------------------------------------------------------------
    # Formatting - laying out background
    # --------------------------------------------------------------------------
    if 'reverse_cols' in formatting:
        for col in formatting['reverse_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "white_on_black")

    if 'reverse_rows' in formatting:
        for rownum in formatting['reverse_rows']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "white_on_black, bold")

    if 'gutter_cols' in formatting:
        for col in formatting['gutter_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "black_on_black")

    if 'gutter_rows' in formatting:
        for rownum in formatting['gutter_rows']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "black_on_black")
            ws.row_dimensions[rownum].height = 5

    if 'conditional_reverse_cols' in formatting:
        for cond_set in formatting['conditional_reverse_cols']:
            for col in cond_set['cols']:
                for i, row in enumerate(data):
                    cond_cell = cond_set['cond'] + str(i+1)

                    if ws[cond_cell].value and 'value' in cond_set:
                        if ws[cond_cell].value == cond_set['value']:
                            cell = ws[col + str(i+1)]
                            apply_formatting_to_cell(cell, "white_on_black")

                    elif ws[cond_cell].value and 'is_number' in cond_set:
                        if cond_set['is_number']:
                            if is_number(ws[cond_cell].value):
                                cell = ws[col + str(i+1)]
                                apply_formatting_to_cell(cell, "white_on_black")
                        else:
                            if not is_number(ws[cond_cell].value):
                                cell = ws[col + str(i+1)]
                                apply_formatting_to_cell(cell, "white_on_black")

                    else:
                        if ws[cond_cell].value:
                            cell = ws[col + str(i+1)]
                            apply_formatting_to_cell(cell, "white_on_black")

    # --------------------------------------------------------------------------
    # Formatting - alternate font for showing code
    # --------------------------------------------------------------------------
    if 'display_columns_as_code' in formatting:
        for col in formatting['display_columns_as_code']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "as_code")

    # --------------------------------------------------------------------------
    # Formatting - Alignment situations
    # --------------------------------------------------------------------------
    if 'wrap_rows' in formatting:
        for rownum in formatting['wrap_rows']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "wrap_on")

    if 'shrink_rows' in formatting:
        for rownum in formatting['shrink_rows']:
            for i, col in enumerate(data[0]):
                cell = ws.cell(column = i+1 , row = rownum )
                apply_formatting_to_cell(cell, "shrink_on")

    if 'shrink_cols' in formatting:
        for col in formatting['shrink_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "shrink_on")

    if 'wrap_cols' in formatting:
        for col in formatting['wrap_cols']:
            for i, row in enumerate(data):
                cell = ws[col + str(i+1)]
                apply_formatting_to_cell(cell, "wrap_on")

    # --------------------------------------------------------------------------
    # Cell-based overrides for formatting user inputs - these will override
    # the reverse-style formatting, as these are generally for props in the
    # heading area
    # --------------------------------------------------------------------------
    if 'user_cells' in formatting:
        for cell in formatting['user_cells']:
            apply_formatting_to_cell(ws[cell], "tint_gray")

    # --------------------------------------------------------------------------
    # Adding hyperlinks
    # --------------------------------------------------------------------------
    if 'hyperlinks' in formatting:
        for list_of_links in formatting['hyperlinks']:
            for link in list_of_links:
                link_from = link[0]
                link_to   = link[1]
                if not link_to.startswith("#"):
                    link_syntax = "#" + wsname + "!" + link_to
                else:
                    link_syntax = link_to
                #print(link)
                ws[link_from].hyperlink = (link_syntax)

    # --------------------------------------------------------------------------
    # Freeze panes
    # --------------------------------------------------------------------------
    if 'freeze_panes' in formatting:
        ws.freeze_panes = formatting['freeze_panes']

    # --------------------------------------------------------------------------
    # Zoom
    # --------------------------------------------------------------------------
    if 'zoom' in formatting:
        ws.sheet_view.zoomScale = formatting['zoom']

    # --------------------------------------------------------------------------
    # Apply dropdowns as validations
    # --------------------------------------------------------------------------
    if 'dropdowns' in formatting:
        for dropdown in formatting['dropdowns']:
            dv             = DataValidation(type="list", formula1=dropdown['list'], allow_blank=dropdown['allow_blank'])
            dv.error       = 'Your entry is not in the list'
            dv.errorTitle  = 'Invalid Entry'
            dv.prompt      = 'Please select from the list'
            dv.promptTitle = 'List Selection'

            dv.add(dropdown['range'])
            #print(dv)

            ws.add_data_validation(dv)

    # --------------------------------------------------------------------------
    # Add comment boxes
    # --------------------------------------------------------------------------
    if 'comments' in formatting:
        for comment_data in formatting['comments']:
            cells  = comment_data[0]
            height = comment_data[1]
            width  = comment_data[2]
            author = comment_data[3]
            text   = comment_data[4]

            comment = xl.comments.Comment(text, author)
            comment.width  = width
            comment.height = height

            for cell in cells:
                ws[cell].comment = comment

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return ws


# ------------------------------------------------------------------------------
# Use some custom notation to make common formatting changes
# ------------------------------------------------------------------------------
def apply_formatting_to_cell(cell, fmt_spec):
    # --------------------------------------------------------------------------
    # Skip out if there's nothing
    # --------------------------------------------------------------------------
    if not fmt_spec:
        return False

    # --------------------------------------------------------------------------
    # Comma separated specs if it's there
    # --------------------------------------------------------------------------
    specs = fmt_spec.split(",")

    # --------------------------------------------------------------------------
    # Fonts 
    # --------------------------------------------------------------------------
    apply_font = False
    fontspecs = "xl.styles.Font("
    for spec in specs:
        spec = spec.strip()
        if 'white_on_black' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "color = 'FFFFFFFF'"
            apply_font = True

        if 'black_on_black' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "color = 'FF000000'"
            apply_font = True

        if 'tint_yellow' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "color = 'FF000000'"
            apply_font = True

        if 'tint_green' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "color = 'FF000000'"
            apply_font = True

        if 'tint_gray' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "color = 'FF000000'"
            apply_font = True

        if 'bold' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "bold = True"
            apply_font = True

        if 'size' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "size = " + str(spec.replace("size=",""))
            apply_font = True

        if 'as_code' in spec:
            if not fontspecs[-1] == "(":
                fontspecs = fontspecs + " , "
            fontspecs = fontspecs + "name = 'Courier', size=9"
            apply_font = True

    fontspecs = fontspecs + ")"

    if apply_font:
        cell.font = eval(fontspecs)

    # --------------------------------------------------------------------------
    # Fill patterns
    # --------------------------------------------------------------------------
    if 'white_on_black' in specs:
        cell.fill = xl.styles.PatternFill(patternType='solid', start_color="FF000000")
    if 'black_on_black' in specs:
        cell.fill = xl.styles.PatternFill(patternType='solid', start_color="FF000000")
    elif 'tint_gray' in specs:
        cell.fill = xl.styles.PatternFill(patternType='solid', start_color="FFd3d3d3")
    elif 'tint_yellow' in specs:
        cell.fill = xl.styles.PatternFill(patternType='solid', start_color="FFFFFF00")
    elif 'tint_green' in specs:
        cell.fill = xl.styles.PatternFill(patternType='solid', start_color="FF00FF00")

    # --------------------------------------------------------------------------
    # Wrap text
    # --------------------------------------------------------------------------
    if 'wrap_on' in specs:
        cell.alignment = xl.styles.Alignment(wrap_text=True)

    # --------------------------------------------------------------------------
    # Shrink text to fit
    # --------------------------------------------------------------------------
    if 'shrink_on' in specs:
        cell.alignment = xl.styles.Alignment(shrink_to_fit=True)

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True


# ------------------------------------------------------------------------------
# 
# ------------------------------------------------------------------------------
def replace_excel_sheet(book, sheet_name, data, fmt):
    tries = 0
    while not can_write_to_excel(book):
        if tries > 5:
            kill_program("File is not available for writing because of permissions issue. Do you have it open?")
        sleep(3)
        tries += 1

    if os.path.isfile(book):
        wb = xl.load_workbook(book)
    else:
        wb = xl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)


    o = write_excel_sheet(wb, sheet_name, data, fmt)
    wb.save(book)
    wb.close()
    return True

